import csv
from openpyxl import Workbook
from openpyxl.styles import Font

ts = list(csv.DictReader(open('data_v2/throughput_summary.csv')))
air = list(csv.DictReader(open('data_v2/tdd_airtime.csv')))


def get(rd, loc, prot, direc, variant=''):
    for r in ts:
        if (r['round'] == rd and r['location'] == loc and r['protocol'] == prot
                and r['direction'] == direc and r['variant'] == variant
                and r['tech'] == '5G'):
            return r
    return None


wb = Workbook()
B = Font(bold=True)
I = Font(italic=True, size=9)

# ---------- Sheet 1: slope chart ----------
s = wb.active
s.title = 'Slope'
s['A1'] = 'Slope chart: TCP throughput, 3:1 vs 2:2 TDD (rx_gain 70 in both rounds)'
s['A1'].font = B
s['A2'] = 'Chart this block: line chart with markers, categories A5:A6, four series B:E'
s['A2'].font = I

series = [('A DL', 'A', 'DL'), ('A UL', 'A', 'UL'),
          ('C DL', 'C', 'DL'), ('C UL', 'C', 'UL')]
s.append([])
s.append(['TDD'] + [n for n, _, _ in series])
for c in range(1, 6):
    s.cell(4, c).font = B
for rd, tag in [('1', '3:1'), ('2', '2:2')]:
    row = [tag]
    for _, loc, direc in series:
        row.append(float(get(rd, loc, 'TCP', direc)['mean_Mbps']))
    s.append(row)

s.append([])
s.append(['Optional error-bar offsets (2:2 row only; the 3:1 round is n=1)'])
s.cell(s.max_row, 1).font = B
s.append(['series', 'mean', 'err_minus', 'err_plus', 'n'])
for c in range(1, 6):
    s.cell(s.max_row, c).font = B
for name, loc, direc in series:
    r = get('2', loc, 'TCP', direc)
    m, lo, hi = float(r['mean_Mbps']), float(r['min_Mbps']), float(r['max_Mbps'])
    s.append([name, m, round(m - lo, 3), round(hi - m, 3), int(r['n_valid'])])
s.column_dimensions['A'].width = 36

# ---------- Sheet 2: slot occupancy ----------
s2 = wb.create_sheet('SlotMap')
s2['A1'] = 'Measured PUSCH occupancy per TDD slot - Location C, rx_gain 60, same session'
s2['A1'].font = B
s2['A2'] = 'Chart this block: clustered column, categories A5:A9, two series B:C'
s2['A2'].font = I
u22 = next(r for r in air if r['unit_id'] == 'R3_5G_C')
u31 = next(r for r in air if r['unit_id'] == 'R3_5G_C_TDD31')
s2.append([])
s2.append(['Slot', '2:2 PUSCH (PRB)', '3:1 PUSCH (PRB)'])
for c in range(1, 4):
    s2.cell(4, c).font = B
for i in range(5):
    s2.append(['Slot ' + str(i),
               float(u22['pusch_rb_slot' + str(i)]),
               float(u31['pusch_rb_slot' + str(i)])])

s2.append([])
s2.append(['Full detail incl. PDSCH - read Notes before charting the PDSCH columns'])
s2.cell(s2.max_row, 1).font = B
s2.append(['Slot', '2:2 PUSCH', '2:2 PDSCH', '3:1 PUSCH', '3:1 PDSCH'])
for c in range(1, 6):
    s2.cell(s2.max_row, c).font = B
for i in range(5):
    s2.append(['Slot ' + str(i),
               float(u22['pusch_rb_slot' + str(i)]),
               float(u22['pdsch_rb_slot' + str(i)]),
               float(u31['pusch_rb_slot' + str(i)]),
               float(u31['pdsch_rb_slot' + str(i)])])
s2.column_dimensions['A'].width = 22
for col in 'BCDE':
    s2.column_dimensions[col].width = 16

# ---------- Sheet 3: reference numbers ----------
s3 = wb.create_sheet('Reference')
s3.append(['Observed change 3:1 -> 2:2 against what airtime alone predicts'])
s3.cell(1, 1).font = B
s3.append(['Airtime prediction from measured slot counts: DL x0.75 (4->3 slots), '
           'UL x2.0 (1->2 slots)'])
s3.cell(2, 1).font = I
s3.append([])
s3.append(['location', 'direction', '3:1 Mb/s', '2:2 Mb/s', 'observed factor',
           'airtime predicts'])
for c in range(1, 7):
    s3.cell(4, c).font = B
for loc in ['A', 'C']:
    for direc in ['DL', 'UL']:
        a = get('1', loc, 'TCP', direc)
        b = get('2', loc, 'TCP', direc)
        m1, m2 = float(a['mean_Mbps']), float(b['mean_Mbps'])
        s3.append([loc, direc, m1, m2, round(m2 / m1, 2),
                   0.75 if direc == 'DL' else 2.0])

s3.append([])
s3.append(['CONFOUND: uplink SNR also moved between campaigns at identical gain settings'])
s3.cell(s3.max_row, 1).font = B
s3.append(['location', 'protocol', 'direction', 'R1 SNR dB', 'R2 SNR dB', 'delta dB'])
for c in range(1, 7):
    s3.cell(s3.max_row, c).font = B
for loc in ['A', 'C']:
    for p in ['TCP', 'UDP']:
        for direc in ['UL', 'DL']:
            a = get('1', loc, p, direc)
            b = get('2', loc, p, direc)
            v1, v2 = float(a['pusch_snr_db']), float(b['pusch_snr_db'])
            s3.append([loc, p, direc, round(v1, 1), round(v2, 1), round(v2 - v1, 1)])
for col in 'ABCDEF':
    s3.column_dimensions[col].width = 16

# ---------- Sheet 4: notes ----------
s4 = wb.create_sheet('Notes')
notes = [
    ('Slope sheet', ''),
    ('', 'Location B is absent on purpose: Round 2 at B is the unusable metal-stand data.'),
    ('', 'The 3:1 row is Round 1, n=1 per condition, so it has no spread. Only the 2:2 row'),
    ('', 'can carry error bars. Offsets are already mean-relative - do not recompute them.'),
    ('', 'DL/UL ratio goes 5.84 -> 1.32 at A (converges) and 17.55 -> 0.39 at C, where'),
    ('', 'uplink ends up 2.6x downlink. That reversal is the headline, not the convergence.'),
    ('', ''),
    ('SlotMap sheet', ''),
    ('', 'The PUSCH columns are the structural claim: uplink can only be scheduled in slot 4'),
    ('', 'under 3:1, and in slots 3-4 under 2:2. That holds whatever traffic was running.'),
    ('', ''),
    ('', 'Use the slot POSITIONS, not the PRB magnitudes. The two units carried different'),
    ('', 'traffic mixes - the 2:2 unit aggregates every test at that location, including'),
    ('', 'downlink runs where the uplink only carries ACKs, while the 3:1 unit is uplink-only.'),
    ('', 'So totalling the PRBs (58.5 at 2:2 vs 42.1 at 3:1, x1.39) is NOT a capacity ratio'),
    ('', 'and must not be quoted as one. The 3:1 slot is fuller (42.1 of 51 PRB) partly'),
    ('', 'because the scheduler compensates, and partly because that run was uplink-only.'),
    ('', ''),
    ('', 'WARNING on the PDSCH columns: the 3:1 unit (R3_5G_C_TDD31) was a TCP UPLINK-ONLY'),
    ('', 'run, so its PDSCH values are near zero because there was no downlink traffic, NOT'),
    ('', 'because the slots were unavailable. Do not chart 3:1 PDSCH against 2:2 PDSCH as if'),
    ('', 'it showed a structural difference. Chart PUSCH only, or state the caveat.'),
    ('', ''),
    ('', 'Round 1 has no airtime data at all - its logs carry no Scheduler-cell METRICS'),
    ('', 'lines. That is why the slot map uses the Round 3 pair (same location, same gain,'),
    ('', 'same session) rather than the Round 1 vs Round 2 pair used on the Slope sheet.'),
    ('', ''),
    ('Reference sheet', ''),
    ('', 'At A the observed factors (DL x0.56, UL x2.48) sit near the airtime prediction.'),
    ('', 'At C the uplink gain is x19.9, an order of magnitude past it - because uplink SNR'),
    ('', 'also rose 22.4 dB between the campaigns at identical gain settings. Caption the'),
    ('', 'figure as the change observed between rounds, not as the effect of the TDD pattern.'),
]
for a, b in notes:
    s4.append([a, b])
    if a:
        s4.cell(s4.max_row, 1).font = B
s4.column_dimensions['A'].width = 18
s4.column_dimensions['B'].width = 96

wb.save('data_v2/fig_tdd_tradeoff.xlsx')
print('wrote data_v2/fig_tdd_tradeoff.xlsx')
print('sheets:', wb.sheetnames)
