"""Export data_v2/ as one formatted workbook.

    python -m pipeline.to_excel

CSV is plain text and carries no formatting: frozen panes, fill colours and
borders cannot exist in a `.csv`. This writes `data_v2/tables.xlsx` instead,
one sheet per table, with:

  * the header row frozen, so it stays visible while scrolling
  * a grey, bold header with an autofilter
  * a black rule under the last row of each round

The CSVs remain the canonical output -- this workbook is for reading, and is
regenerated from them. Requires openpyxl; the rest of the pipeline does not.
"""
from __future__ import annotations

import csv
import sys

from . import paths

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is needed for the Excel export:\n    python -m pip install openpyxl")

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(bold=True)
RULE = Border(bottom=Side(style="medium", color="000000"))

# sheet name -> column whose changes draw a rule. Tables already sort by round.
GROUP_BY = {
    "reps": "round", "throughput_summary": "round", "phy_by_unit": "round",
    "events_by_unit": "round", "logging_cost": "round", "tdd_airtime": "round",
    "realtime": "round", "ping_series": "location", "ping_summary": "location",
}

# read as text, not numbers -- Excel would reinterpret these
TEXT_COLS = {"ts", "start_utc", "rep_id", "unit_id", "campaign", "tdd",
             "throughput_src", "rnti", "variant"}


def _num(v):
    """Convert to a number where that is what it is, else leave as text."""
    if v is None or v == "":
        return None
    try:
        f = float(v)
    except ValueError:
        return v
    return int(f) if f.is_integer() and abs(f) < 1e15 else f


def add_sheet(wb, name, rows, header):
    ws = wb.create_sheet(name[:31])
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    key = GROUP_BY.get(name)
    for r in rows:
        ws.append([r.get(h) if h in TEXT_COLS else _num(r.get(h)) for h in header])

    # black rule under the last row of each group
    if key and key in header:
        for i, r in enumerate(rows):
            nxt = rows[i + 1] if i + 1 < len(rows) else None
            if nxt is None or nxt.get(key) != r.get(key):
                if nxt is None:
                    continue                      # no rule under the final row
                for c in range(1, len(header) + 1):
                    ws.cell(row=i + 2, column=c).border = RULE

    ws.freeze_panes = "A2"                        # keep the header visible
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows)+1}"
    ws.row_dimensions[1].height = 30

    for i, h in enumerate(header, start=1):
        widest = max([len(str(h))] + [len(str(r.get(h) or "")) for r in rows[:400]])
        ws.column_dimensions[get_column_letter(i)].width = min(max(widest + 2, 9), 26)
    return ws


def main():
    files = sorted(paths.DATA_V2.glob("*.csv"))
    if not files:
        sys.exit(f"no CSVs in {paths.DATA_V2} -- run `python -m pipeline.build_tables` first")

    wb = Workbook()
    wb.remove(wb.active)
    # most-used tables first
    order = ["reps", "throughput_summary", "phy_by_unit", "tdd_airtime",
             "realtime", "events_by_unit", "logging_cost", "ping_summary",
             "ping_series"]
    files.sort(key=lambda p: (order.index(p.stem) if p.stem in order else 99, p.stem))

    for p in files:
        with open(p, newline="", encoding="utf-8") as f:
            rd = csv.DictReader(f)
            rows, header = list(rd), rd.fieldnames or []
        if not header:
            continue
        add_sheet(wb, p.stem, rows, header)
        print(f"  {p.stem:24s} {len(rows):5d} rows x {len(header):2d} cols")

    out = paths.DATA_V2 / "tables.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        sys.exit(f"{out.name} is open in Excel -- close it and re-run")
    print(f"-> {out}")


if __name__ == "__main__":
    main()
